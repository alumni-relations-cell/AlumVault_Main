"""
AlumVault Excel importer (CLI).

Reads an .xlsx file matching the template at
scripts/templates/alumni_import_template.xlsx, dedups within the sheet and
against the existing DB, and writes results to Postgres.

Usage:
    python scripts/import/import_excel.py path/to/sheet.xlsx
    python scripts/import/import_excel.py path/to/sheet.xlsx --dry-run
"""
from __future__ import annotations
import sys
import os
import argparse
from pathlib import Path
from collections import defaultdict

# Make sibling modules importable when invoked from project root
sys.path.insert(0, str(Path(__file__).parent))

import psycopg2
import psycopg2.extras
from openpyxl import load_workbook
from dotenv import load_dotenv

from normalizer import normalize_row, blind_index

# Load .env from project root
ROOT = Path(__file__).resolve().parent.parent.parent
load_dotenv(ROOT / ".env")

# Matcher score weights — must match the Go Matcher's logic in spirit
SCORE_LINKEDIN = 100
SCORE_EMAIL    = 90
SCORE_PHONE    = 60
SCORE_NAME     = 40
SCORE_BATCH    = 20
SCORE_BRANCH   = 20
SCORE_COMPANY  = 10

AUTO_MERGE_THRESHOLD = 100
REVIEW_THRESHOLD     = 70


def db_connect():
    return psycopg2.connect(
        host=os.environ["DB_HOST"],
        port=int(os.environ["DB_PORT"]),
        user=os.environ["DB_USER"],
        password=os.environ["DB_PASSWORD"],
        dbname=os.environ["DB_NAME"],
    )


def read_sheet(path: Path) -> list[dict]:
    """Read the alumni_data sheet (or first sheet) into a list of dicts."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["alumni_data"] if "alumni_data" in wb.sheetnames else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h else "" for h in next(rows_iter)]

    rows = []
    for r in rows_iter:
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        rows.append(dict(zip(headers, r)))
    return rows


def dedup_within_sheet(rows: list[dict], blind_key: str) -> list[dict]:
    """
    Collapse same-sheet duplicates BEFORE hitting the DB.
    Match keys (in order of strength):
        1. linkedin_url
        2. any shared email
        3. (full_name_blind, batch_year, branch)
    Merging: union for emails/phones/companies/tags; newest non-empty wins for scalars.
    """
    merged: list[dict] = []
    by_linkedin: dict[str, int] = {}
    by_email: dict[str, int] = {}
    by_blind: dict[tuple, int] = {}

    for raw in rows:
        n = normalize_row(raw)
        if not n["full_name"]:
            continue  # rows without a name are skipped entirely
        n["full_name_blind"] = blind_index(n["full_name"], blind_key)

        match_idx = None
        if n["linkedin_url"] and n["linkedin_url"] in by_linkedin:
            match_idx = by_linkedin[n["linkedin_url"]]
        if match_idx is None:
            for e in n["emails"]:
                if e in by_email:
                    match_idx = by_email[e]
                    break
        if match_idx is None and n["full_name_blind"] and n["batch_year"] and n["branch"]:
            key = (n["full_name_blind"], n["batch_year"], n["branch"])
            if key in by_blind:
                match_idx = by_blind[key]

        if match_idx is None:
            merged.append(n)
            idx = len(merged) - 1
            if n["linkedin_url"]:
                by_linkedin[n["linkedin_url"]] = idx
            for e in n["emails"]:
                by_email.setdefault(e, idx)
            if n["full_name_blind"] and n["batch_year"] and n["branch"]:
                by_blind[(n["full_name_blind"], n["batch_year"], n["branch"])] = idx
        else:
            _merge_into(merged[match_idx], n)
            # newly-revealed match keys also get indexed against the same record
            if n["linkedin_url"]:
                by_linkedin.setdefault(n["linkedin_url"], match_idx)
            for e in n["emails"]:
                by_email.setdefault(e, match_idx)

    return merged


def _merge_into(target: dict, source: dict):
    """Merge source row into target in-place. Union lists, last-non-empty for scalars."""
    LIST_FIELDS = ("emails", "phones", "tags")
    for f in LIST_FIELDS:
        seen = {x.lower() if isinstance(x, str) else x for x in target.get(f, [])}
        for v in source.get(f, []):
            key = v.lower() if isinstance(v, str) else v
            if key not in seen:
                target[f].append(v)
                seen.add(key)

    # Companies: union by lowercased name, latest is_current=True wins
    existing = {c["company"].lower(): c for c in target.get("companies", [])}
    for c in source.get("companies", []):
        k = c["company"].lower()
        if k in existing:
            if c["is_current"]:
                existing[k]["is_current"] = True
        else:
            existing[k] = c
    target["companies"] = list(existing.values())

    # Scalars: prefer source if source is non-empty
    SCALARS = (
        "linkedin_url", "batch_year", "degree", "branch", "enrollment_no", "dob",
        "current_company", "current_title", "industry", "current_city",
        "current_country", "notes", "full_name_blind",
    )
    for f in SCALARS:
        if source.get(f) and not target.get(f):
            target[f] = source[f]


def score_against_db(cur, candidate: dict) -> tuple[int, str | None]:
    """
    Find the best-matching existing alumni row for this candidate.
    Returns (score, alumni_id_or_None).
    """
    best_score = 0
    best_id = None

    # 1. LinkedIn URL exact match (strongest signal)
    if candidate["linkedin_url"]:
        cur.execute(
            "SELECT id FROM alumni WHERE LOWER(linkedin_url) = LOWER(%s) LIMIT 1",
            (candidate["linkedin_url"],),
        )
        row = cur.fetchone()
        if row:
            return SCORE_LINKEDIN, row[0]  # short-circuit — definitive match

    # 2. Email match — look in alumni.emails JSONB
    if candidate["emails"]:
        cur.execute(
            """
            SELECT id FROM alumni
             WHERE emails ?| %s
             LIMIT 1
            """,
            (candidate["emails"],),
        )
        row = cur.fetchone()
        if row:
            return SCORE_EMAIL, row[0]

    # 3. Phone match
    if candidate["phones"]:
        cur.execute(
            """
            SELECT id FROM alumni
             WHERE phones ?| %s
             LIMIT 1
            """,
            (candidate["phones"],),
        )
        row = cur.fetchone()
        if row:
            return SCORE_PHONE, row[0]

    # 4. Name + batch + branch via blind index — broader candidate sweep
    if candidate["full_name_blind"]:
        cur.execute(
            """
            SELECT id, batch_year, branch, current_company
              FROM alumni
             WHERE full_name_blind = %s
             LIMIT 20
            """,
            (candidate["full_name_blind"],),
        )
        for row in cur.fetchall():
            score = SCORE_NAME
            if candidate["batch_year"] and row[1] == candidate["batch_year"]:
                score += SCORE_BATCH
            if candidate["branch"] and row[2] == candidate["branch"]:
                score += SCORE_BRANCH
            if (candidate["current_company"] and row[3]
                    and candidate["current_company"].lower() == row[3].lower()):
                score += SCORE_COMPANY
            if score > best_score:
                best_score = score
                best_id = row[0]

    return best_score, best_id


def insert_new_alumni(cur, candidate: dict) -> str:
    field_sources = {
        f: candidate["source_label"]
        for f in ("full_name", "linkedin_url", "emails", "phones",
                  "batch_year", "degree", "branch", "current_company",
                  "current_title", "current_city")
        if candidate.get(f)
    }

    cur.execute(
        """
        INSERT INTO alumni (
            full_name, full_name_blind, enrollment_no, batch_year, branch, degree, dob,
            emails, phones, current_company, current_title, industry,
            linkedin_url, current_city, field_sources, tags,
            data_completeness, overall_confidence
        )
        VALUES (
            %(full_name)s, %(full_name_blind)s, %(enrollment_no)s, %(batch_year)s,
            %(branch)s, %(degree)s, %(dob)s,
            %(emails)s, %(phones)s, %(current_company)s, %(current_title)s,
            %(industry)s, %(linkedin_url)s, %(current_city)s,
            %(field_sources)s, %(tags)s,
            %(completeness)s, 0.7
        )
        RETURNING id
        """,
        {
            **candidate,
            "emails": psycopg2.extras.Json(candidate["emails"]),
            "phones": psycopg2.extras.Json(candidate["phones"]),
            "field_sources": psycopg2.extras.Json(field_sources),
            "completeness": _completeness(candidate),
        },
    )
    alumni_id = cur.fetchone()[0]

    for company in candidate["companies"]:
        cur.execute(
            """
            INSERT INTO alumni_companies (alumni_id, company, title, is_current, source)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (alumni_id, company) DO NOTHING
            """,
            (
                alumni_id, company["company"],
                candidate["current_title"] if company["is_current"] else None,
                company["is_current"], candidate["source_label"],
            ),
        )

    return alumni_id


def merge_into_existing(cur, alumni_id: str, candidate: dict):
    """Union emails/phones/companies into existing alumni row; fill in missing scalars."""
    cur.execute(
        """
        UPDATE alumni
           SET emails = (
                   SELECT to_jsonb(array(SELECT DISTINCT jsonb_array_elements_text(
                       COALESCE(emails, '[]'::jsonb) || %s::jsonb
                   )))
               ),
               phones = (
                   SELECT to_jsonb(array(SELECT DISTINCT jsonb_array_elements_text(
                       COALESCE(phones, '[]'::jsonb) || %s::jsonb
                   )))
               ),
               linkedin_url    = COALESCE(linkedin_url, %s),
               current_company = COALESCE(NULLIF(current_company, ''), %s),
               current_title   = COALESCE(NULLIF(current_title, ''), %s),
               current_city    = COALESCE(NULLIF(current_city, ''), %s),
               industry        = COALESCE(NULLIF(industry, ''), %s),
               enrollment_no   = COALESCE(NULLIF(enrollment_no, ''), %s),
               dob             = COALESCE(dob, %s),
               degree          = COALESCE(NULLIF(degree, ''), %s),
               branch          = COALESCE(NULLIF(branch, ''), %s),
               batch_year      = COALESCE(batch_year, %s),
               tags            = ARRAY(SELECT DISTINCT unnest(COALESCE(tags, '{}') || %s::text[])),
               updated_at      = NOW()
         WHERE id = %s
        """,
        (
            psycopg2.extras.Json(candidate["emails"]),
            psycopg2.extras.Json(candidate["phones"]),
            candidate["linkedin_url"],
            candidate["current_company"], candidate["current_title"],
            candidate["current_city"], candidate["industry"],
            candidate["enrollment_no"], candidate["dob"],
            candidate["degree"], candidate["branch"],
            candidate["batch_year"], candidate["tags"], alumni_id,
        ),
    )

    for company in candidate["companies"]:
        cur.execute(
            """
            INSERT INTO alumni_companies (alumni_id, company, title, is_current, source)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (alumni_id, company)
            DO UPDATE SET is_current = alumni_companies.is_current OR EXCLUDED.is_current
            """,
            (
                alumni_id, company["company"],
                candidate["current_title"] if company["is_current"] else None,
                company["is_current"], candidate["source_label"],
            ),
        )


def queue_for_review(cur, candidate: dict, existing_alumni_id: str, score: int):
    """Insert a low-confidence match into review_queue for a human to resolve."""
    breakdown = {"total": score, "matcher": "excel_importer_v1"}
    cur.execute(
        """
        INSERT INTO review_queue (
            existing_alumni_id, incoming_data, match_score, score_breakdown, status
        )
        VALUES (%s, %s::jsonb, %s, %s::jsonb, 'pending')
        """,
        (
            existing_alumni_id,
            psycopg2.extras.Json({
                **candidate,
                "dob": str(candidate["dob"]) if candidate["dob"] else None,
            }),
            score,
            psycopg2.extras.Json(breakdown),
        ),
    )


def _completeness(c: dict) -> float:
    """Fraction of important fields that have data. 0.0 - 1.0."""
    important = ("linkedin_url", "emails", "phones", "batch_year", "branch",
                 "current_company", "current_city")
    have = sum(1 for f in important if c.get(f))
    return round(have / len(important), 2)


def main():
    parser = argparse.ArgumentParser(description="AlumVault Excel importer")
    parser.add_argument("file", type=Path, help="Path to .xlsx file matching the template")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print results without writing to DB")
    args = parser.parse_args()

    if not args.file.exists():
        sys.exit(f"File not found: {args.file}")

    blind_key = os.environ["BLIND_INDEX_KEY"]

    print(f"Reading {args.file}")
    raw_rows = read_sheet(args.file)
    print(f"   {len(raw_rows)} raw rows")

    candidates = dedup_within_sheet(raw_rows, blind_key)
    print(f"   {len(candidates)} unique candidates after within-sheet dedup "
          f"({len(raw_rows) - len(candidates)} merged)")

    if args.dry_run:
        print("\nDRY RUN - sample of normalized candidates:")
        for c in candidates[:5]:
            print(f"  - {c['full_name']}  linkedin={c['linkedin_url']}  "
                  f"emails={c['emails']}  batch={c['batch_year']}  branch={c['branch']}")
        print(f"\n  ... ({len(candidates)} total) - DB not modified.")
        return

    conn = db_connect()
    conn.autocommit = False
    cur = conn.cursor()

    stats = defaultdict(int)
    review_rows = []

    try:
        for c in candidates:
            score, alumni_id = score_against_db(cur, c)
            if alumni_id and score >= AUTO_MERGE_THRESHOLD:
                merge_into_existing(cur, alumni_id, c)
                stats["merged"] += 1
            elif alumni_id and score >= REVIEW_THRESHOLD:
                queue_for_review(cur, c, alumni_id, score)
                stats["queued_for_review"] += 1
                review_rows.append((c["full_name"], score))
            else:
                insert_new_alumni(cur, c)
                stats["inserted"] += 1

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()

    print("\nImport complete")
    print(f"   inserted          : {stats['inserted']}")
    print(f"   merged (auto)     : {stats['merged']}")
    print(f"   queued for review : {stats['queued_for_review']}")
    if review_rows:
        print("\n   Review queue (top 10):")
        for name, score in review_rows[:10]:
            print(f"     - {name}  (score={score})")


if __name__ == "__main__":
    main()
