"""
Generates the canonical AlumVault Excel import template.
Run:  python scripts/templates/generate_import_template.py
Output: scripts/templates/alumni_import_template.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# (column_name, required, description, example)
COLUMNS = [
    ("full_name",        "YES",     "Alumnus full name. Trim spaces. Title case fine.",            "Rahul Sharma"),
    ("linkedin_url",     "STRONG",  "Best dedup key. Any format ok — we normalize.",                "https://linkedin.com/in/rahul-sharma-1234"),
    ("email",            "",        "Primary email. Lowercased on import.",                        "rahul.sharma@gmail.com"),
    ("email_alt",        "",        "Optional secondary email. Add email_alt2, email_alt3 if needed.", "rahul@company.com"),
    ("phone",            "",        "Any format — we strip +91, spaces, dashes.",                  "+91 98765 43210"),
    ("phone_alt",        "",        "Optional secondary phone.",                                   "98765 43211"),
    ("batch_year",       "",        "Year of graduation, e.g. 2018",                               "2018"),
    ("degree",           "",        "B.Tech / M.Tech / Ph.D / MBA / etc.",                         "B.Tech"),
    ("branch",           "",        "CSE / ECE / ME / EE / etc — synonyms normalized.",            "CSE"),
    ("enrollment_no",    "",        "Thapar enrollment number, if available.",                     "101803123"),
    ("dob",              "",        "Date of birth — YYYY-MM-DD or DD/MM/YYYY.",                    "1996-04-15"),
    ("current_company",  "",        "Latest known employer. Stored as is_current=true.",           "Google"),
    ("current_title",    "",        "Current role / designation.",                                 "Senior Software Engineer"),
    ("past_companies",   "",        "Pipe-separated history: 'Microsoft|Amazon|Flipkart'",         "Microsoft|Amazon"),
    ("industry",         "",        "Industry / sector.",                                          "Software"),
    ("current_city",     "",        "City — 'Bangalore', 'Mumbai', 'San Francisco'.",              "Bangalore"),
    ("current_country",  "",        "Country.",                                                    "India"),
    ("user_type",        "",        "alumnus / student / faculty — pushed to tags.",                "alumnus"),
    ("tags",             "",        "Pipe-separated extra tags: 'mentor|donor|startup_founder'.",  "mentor|startup_founder"),
    ("notes",            "",        "Free text — anything else worth recording.",                  "Spoke at TechFest 2023"),
    ("source_label",     "",        "Where this row came from. Used in field_sources audit.",      "batch_2018_sheet"),
]

# Color palette
HEADER_FILL    = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
REQUIRED_FILL  = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
STRONG_FILL    = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
SAMPLE_FILL    = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ZEBRA_FILL     = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

HEADER_FONT    = Font(bold=True, color="FFFFFF", size=11)
NORMAL_FONT    = Font(size=10)
BOLD_FONT      = Font(bold=True, size=11)

THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def autosize(ws, min_w=12, max_w=45):
    """Resize columns based on content length."""
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        ws.column_dimensions[letter].width = max(min_w, min(max_w, max_len + 2))


def build_data_sheet(wb):
    ws = wb.active
    ws.title = "alumni_data"
    ws.freeze_panes = "A2"

    # Header row
    for idx, (name, req, _desc, _ex) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=name)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        if req == "YES":
            cell.fill = REQUIRED_FILL
        elif req == "STRONG":
            cell.fill = STRONG_FILL
        else:
            cell.fill = HEADER_FILL

    # Example row (green tint to make it obvious it's a template, not real data)
    for idx, (_n, _r, _d, example) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=idx, value=example)
        cell.fill = SAMPLE_FILL
        cell.font = NORMAL_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER

    # A second example row — minimal data only (LinkedIn + name)
    sparse = ["Priya Verma", "https://linkedin.com/in/priya-verma-9876"] + [""] * (len(COLUMNS) - 2)
    for idx, val in enumerate(sparse, start=1):
        cell = ws.cell(row=3, column=idx, value=val)
        cell.fill = SAMPLE_FILL
        cell.font = NORMAL_FONT
        cell.border = BORDER

    autosize(ws)
    ws.row_dimensions[1].height = 24
    return ws


def build_readme_sheet(wb):
    ws = wb.create_sheet("README", 0)  # put README first so it opens first
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 35

    intro = [
        ("AlumVault — Excel Import Template",                                         BOLD_FONT),
        ("",                                                                            None),
        ("How to use this template",                                                    BOLD_FONT),
        ("1. Fill rows in the 'alumni_data' sheet — one row per alumnus.",              None),
        ("2. Delete the two example rows (they are highlighted green) before import.",  None),
        ("3. Leave cells blank when you don't have the data — importer treats blanks",  None),
        ("   as 'unknown' (won't overwrite existing values in the database).",          None),
        ("4. If a column doesn't apply to your sheet, you can delete the entire column.", None),
        ("5. Save as .xlsx (or .csv) and upload via Admin → Import.",                   None),
        ("",                                                                            None),
        ("Column requirement legend",                                                   BOLD_FONT),
        ("  RED header     = required",                                                 None),
        ("  ORANGE header  = strongly recommended (best dedup key)",                    None),
        ("  BLUE header    = optional",                                                 None),
        ("",                                                                            None),
        ("How duplicate detection works",                                               BOLD_FONT),
        ("• Within the same sheet:",                                                    None),
        ("    - Same LinkedIn URL → auto-merged (rows combined into one).",             None),
        ("    - Same name + batch_year + branch (no LinkedIn) → auto-merged if",        None),
        ("      no field conflicts; otherwise flagged for human review.",               None),
        ("• Across sheets / against existing database:",                                 None),
        ("    - Weighted scoring: LinkedIn=100, email=90, phone=60, name-fuzzy=40,",    None),
        ("      batch=20, branch=20, company=10.",                                      None),
        ("    - Score >= 100 → auto-merge.",                                            None),
        ("    - Score 70-99 → goes to review queue (you decide).",                      None),
        ("    - Score < 70  → treated as a new person.",                                None),
        ("• Merge rules:",                                                              None),
        ("    - Emails / phones / past_companies → union (all kept, de-duplicated).",   None),
        ("    - Single-value fields → newest non-empty wins; source tracked in",        None),
        ("      field_sources JSONB for audit.",                                        None),
        ("",                                                                            None),
        ("Normalisation applied automatically",                                          BOLD_FONT),
        ("• Names → trimmed, multiple spaces collapsed.",                                None),
        ("• Emails → lowercased.",                                                       None),
        ("• Phones → +91, dashes, spaces stripped; left as 10-digit number.",            None),
        ("• LinkedIn URLs → canonicalized to 'linkedin.com/in/<slug>' (no trailing",     None),
        ("  slash, no query params).",                                                   None),
        ("• Branch synonyms → 'Computer Science', 'CS', 'CSE', 'Computer Science",        None),
        ("  Engineering' all map to canonical 'CSE'.",                                    None),
        ("• 'N/A', '-', '—', 'NA', 'null' → treated as empty.",                          None),
        ("",                                                                            None),
        ("Column reference",                                                             BOLD_FONT),
    ]

    row = 1
    for text, font in intro:
        c = ws.cell(row=row, column=1, value=text)
        if font:
            c.font = font
        if text.startswith("AlumVault —"):
            c.font = Font(bold=True, size=14, color="1F4E78")
        row += 1

    # Column reference table header
    headers = ["Column", "Required", "Description", "Example"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for i, (name, req, desc, example) in enumerate(COLUMNS):
        zebra = (i % 2 == 1)
        fill = ZEBRA_FILL if zebra else None

        cells = [
            (1, name, BOLD_FONT),
            (2, req if req else "optional", None),
            (3, desc, None),
            (4, example, None),
        ]
        for col, val, font in cells:
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if font:
                cell.font = font
            else:
                cell.font = NORMAL_FONT
            if fill:
                cell.fill = fill

        # Color the Required cell
        req_cell = ws.cell(row=row, column=2)
        if req == "YES":
            req_cell.fill = REQUIRED_FILL
            req_cell.font = Font(color="FFFFFF", bold=True, size=10)
        elif req == "STRONG":
            req_cell.fill = STRONG_FILL
            req_cell.font = Font(color="FFFFFF", bold=True, size=10)
        row += 1

    ws.freeze_panes = "A2"
    return ws


def build_branch_synonyms_sheet(wb):
    """Reference sheet showing how branch names get normalized."""
    ws = wb.create_sheet("branch_synonyms")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 70

    headers = ["Canonical", "Synonyms (any of these are accepted)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")

    synonyms = [
        ("CSE",  "Computer Science, Computer Science Engineering, CS, COE, Computer Engg, COMP"),
        ("ECE",  "Electronics, Electronics & Communication, Electronics and Communication, EC, ENC"),
        ("EE",   "Electrical, Electrical Engineering, Electrical Engg"),
        ("EIC",  "Electronics & Instrumentation, Electronics and Instrumentation, Instrumentation"),
        ("ME",   "Mechanical, Mechanical Engineering, Mechanical Engg, MECH"),
        ("CHE",  "Chemical, Chemical Engineering, Chem"),
        ("CIVIL","Civil, Civil Engineering"),
        ("BIO",  "Biotechnology, Biotech, BT, Bio Technology"),
        ("MBA",  "MBA, Master of Business Administration, LMTSOM"),
        ("MCA",  "MCA, Master of Computer Applications"),
        ("BBA",  "BBA, Bachelor of Business Administration"),
    ]
    row = 2
    for canonical, syn in synonyms:
        c1 = ws.cell(row=row, column=1, value=canonical)
        c1.font = BOLD_FONT
        c1.border = BORDER
        c2 = ws.cell(row=row, column=2, value=syn)
        c2.font = NORMAL_FONT
        c2.border = BORDER
        c2.alignment = Alignment(wrap_text=True)
        if row % 2 == 1:
            c1.fill = ZEBRA_FILL
            c2.fill = ZEBRA_FILL
        row += 1

    note = ws.cell(
        row=row + 1, column=1,
        value="NOTE: this list is a starter. Tell the AlumVault team if a Thapar branch is missing or wrong."
    )
    note.font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=2)
    return ws


def main():
    wb = Workbook()
    build_readme_sheet(wb)
    build_data_sheet(wb)
    build_branch_synonyms_sheet(wb)

    out = Path(__file__).parent / "alumni_import_template.xlsx"
    wb.save(out)
    print(f"Generated: {out}")


if __name__ == "__main__":
    main()
